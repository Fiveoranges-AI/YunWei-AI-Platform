"""Spreadsheet parser.

Handles ``.xlsx`` via ``openpyxl``, ``.csv`` via stdlib ``csv``, and
``.xls`` via ``pandas`` (with ``xlrd`` as the optional engine if the
file format requires it). Every non-empty cell becomes its own chunk
with id ``sheet:<sheet_name>!R<row>C<col>`` so extracted fields can
ground back to the exact cell.
"""

from __future__ import annotations

import asyncio
import csv
import os
from pathlib import Path
from typing import Any

from yunwei_win.services.schema_ingest.parse_artifact import (
    ParseArtifact,
    ParseCapabilities,
    ParseChunk,
    ParseTable,
    ParseTableCell,
)


class SpreadsheetParser:
    async def parse_file(
        self,
        path: Path,
        *,
        filename: str,
        content_type: str | None,
        source_type: str = "spreadsheet",
    ) -> ParseArtifact:
        return await asyncio.to_thread(
            self._parse_sync, path, filename, content_type, source_type
        )

    def _parse_sync(
        self,
        path: Path,
        filename: str,
        content_type: str | None,
        source_type: str,
    ) -> ParseArtifact:
        ext = os.path.splitext(filename)[1].lower()
        ct = (content_type or "").lower()

        if ext == ".csv" or ct in {"text/csv", "application/csv"}:
            sheets = _read_csv(path, filename)
        elif ext == ".xls" or ct == "application/vnd.ms-excel":
            sheets = _read_xls(path)
        else:
            sheets = _read_xlsx(path)

        chunks: list[ParseChunk] = []
        tables: list[ParseTable] = []
        grounding: dict[str, Any] = {}
        markdown_parts: list[str] = []

        for sheet_name, rows in sheets:
            cells: list[ParseTableCell] = []
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, value in enumerate(row, start=1):
                    if value is None or value == "":
                        continue
                    text = _to_text(value)
                    cell_id = f"sheet:{sheet_name}!R{r_idx}C{c_idx}"
                    chunks.append(
                        ParseChunk(id=cell_id, type="spreadsheet_cell", text=text)
                    )
                    cells.append(
                        ParseTableCell(
                            row=r_idx, col=c_idx, text=text, ref_id=cell_id
                        )
                    )
                    grounding[cell_id] = {
                        "sheet": sheet_name,
                        "row": r_idx,
                        "col": c_idx,
                    }
            tables.append(ParseTable(id=f"sheet:{sheet_name}", sheet=sheet_name, cells=cells))
            markdown_parts.append(_sheet_to_markdown(sheet_name, rows))

        markdown = "\n\n".join(markdown_parts)

        return ParseArtifact(
            version=1,
            provider="spreadsheet",
            source_type=source_type,
            markdown=markdown,
            pages=[{"id": f"sheet:{name}", "sheet_name": name} for name, _ in sheets],
            chunks=chunks,
            grounding=grounding,
            tables=tables,
            metadata={"filename": filename, "sheet_count": len(sheets)},
            capabilities=ParseCapabilities(
                pages=True,
                chunks=True,
                spreadsheet_cells=True,
                table_cells=True,
            ),
        )


def _read_xlsx(path: Path) -> list[tuple[str, list[list[Any]]]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets.append((sheet_name, rows))
    wb.close()
    return sheets


def _read_xls(path: Path) -> list[tuple[str, list[list[Any]]]]:
    import pandas as pd

    frames = pd.read_excel(str(path), sheet_name=None, header=None)
    sheets: list[tuple[str, list[list[Any]]]] = []
    for sheet_name, df in frames.items():
        rows = df.where(df.notna(), None).values.tolist()
        sheets.append((str(sheet_name), rows))
    return sheets


def _read_csv(path: Path, filename: str) -> list[tuple[str, list[list[Any]]]]:
    sheet_name = os.path.splitext(filename)[0] or "Sheet1"
    rows: list[list[Any]] = []
    with open(path, encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        for row in reader:
            rows.append(list(row))
    return [(sheet_name, rows)]


def _to_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_to_markdown(sheet_name: str, rows: list[list[Any]]) -> str:
    lines = [f"## {sheet_name}"]
    for row in rows:
        if not row:
            continue
        rendered = " | ".join(
            "" if v is None else _to_text(v) for v in row
        )
        lines.append(f"| {rendered} |")
    return "\n".join(lines)
