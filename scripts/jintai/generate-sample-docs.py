"""Generate 3 sample documents for round 5 frontend upload UI demo.

Files land in apps/win-web/public/samples/jintai/:
  * 领料单.jpg     — 模拟手写领料单照片 (Pillow 1024×768)
  * 采购合同.pdf   — minimal PDF (pure Python, no reportlab dep)
  * 供应商对账.xlsx — openpyxl 3 sheets

Content is "演示样本" watermarked / labelled so reviewer 不会混淆生产数据.
DemoMockProvider only looks at filename + size (deterministic seed) so file
contents are advisory; they exist mostly so the user can preview before upload.

Run:
  python3 scripts/jintai/generate-sample-docs.py
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[2]
TARGET = REPO_ROOT / "apps" / "win-web" / "public" / "samples" / "jintai"


def main() -> int:
    TARGET.mkdir(parents=True, exist_ok=True)
    n = 0
    n += _gen_issue_voucher_jpg(TARGET / "领料单.jpg")
    n += _gen_purchase_contract_pdf(TARGET / "采购合同.pdf")
    n += _gen_supplier_reconciliation_xlsx(TARGET / "供应商对账.xlsx")
    print(f"\n生成 {n} 个示例文档 → {TARGET}/")
    return 0


# ============================== 领料单.jpg ==============================


def _gen_issue_voucher_jpg(path: Path) -> int:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("Pillow 缺失,跳过 .jpg", file=sys.stderr)
        return 0

    img = Image.new("RGB", (1024, 768), color="#fcfaf6")
    draw = ImageDraw.Draw(img)

    # 找一个支持中文的字体. macOS 系统字体优先.
    font_candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ]
    title_font = body_font = None
    for fc in font_candidates:
        if Path(fc).exists():
            title_font = ImageFont.truetype(fc, 36)
            body_font = ImageFont.truetype(fc, 22)
            break
    if title_font is None:
        title_font = body_font = ImageFont.load_default()
        print("  (warn: 没找到中文字体,中文将渲染为方块)")

    # 标题
    draw.text((40, 30), "宜兴市锦泰耐火材料 · 车间领料单", fill="#1a1a1a", font=title_font)
    draw.line([(40, 80), (984, 80)], fill="#888", width=2)

    # 表格内容 (模拟手写,但用印刷字体方便 OCR)
    rows = [
        ("领料单号", "BL-2026-018"),
        ("申请车间", "成型车间"),
        ("领用人", "张师傅"),
        ("物料名称", "α 氧化铝粉"),
        ("规格", "CT3000SG · 5N 级"),
        ("数量", "800 kg"),
        ("用途", "BL-2026-018 容百二供 NCM 高镍配料"),
        ("领用日期", str(date.today())),
        ("库管签名", "王仓管 ✓"),
    ]
    y = 120
    for label, value in rows:
        draw.text((60, y), f"{label}:", fill="#444", font=body_font)
        draw.text((240, y), str(value), fill="#1a1a1a", font=body_font)
        y += 48

    # 水印
    draw.text((40, 700), "★ 演示样本 — 锦泰 demo 用,非真实业务文档 ★",
              fill="#c0392b", font=body_font)

    img.save(path, "JPEG", quality=85)
    print(f"  ✓ {path.relative_to(REPO_ROOT)}  ({path.stat().st_size} bytes)")
    return 1


# ============================== 采购合同.pdf ============================


def _gen_purchase_contract_pdf(path: Path) -> int:
    """Minimal PDF 1.4 with English-only text (Helvetica latin)."""
    text_lines = [
        "JinTai Refractories Co.  Purchase Contract  DEMO 2026-Q2",
        "",
        "Vendor:    Shandong Zhonglv Materials",
        "Product:   Alpha-Alumina Powder (CT3000SG 5N grade)",
        "Quantity:  4000 kg",
        "UnitPrice: CNY 24.00 / kg",
        "Total:     CNY 96000.00",
        "Delivery:  2026-06-05",
        "Terms:     60 days",
        "",
        "DEMO SAMPLE - JinTai backend mode demo only.",
    ]
    content_stream_parts = [b"BT /F1 13 Tf 50 780 Td 16 TL"]
    for line in text_lines:
        esc = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content_stream_parts.append(f"({esc}) Tj T*".encode("ascii"))
    content_stream_parts.append(b"ET")
    cb = b"\n".join(content_stream_parts)

    objs: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        b"<< /Length " + str(len(cb)).encode("ascii") + b" >>\nstream\n" + cb + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]

    pdf = b"%PDF-1.4\n%\xc2\xa5\xc2\xb1\xc3\xab\n"
    offsets: list[int] = []
    for i, obj in enumerate(objs, start=1):
        offsets.append(len(pdf))
        pdf += f"{i} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"

    xref_offset = len(pdf)
    pdf += b"xref\n0 6\n0000000000 65535 f \n"
    for o in offsets:
        pdf += f"{o:010d} 00000 n \n".encode("ascii")
    pdf += (
        b"trailer << /Size 6 /Root 1 0 R >>\nstartxref\n"
        + str(xref_offset).encode("ascii")
        + b"\n%%EOF\n"
    )

    path.write_bytes(pdf)
    print(f"  ✓ {path.relative_to(REPO_ROOT)}  ({path.stat().st_size} bytes)")
    return 1


# ============================== 供应商对账.xlsx =========================


def _gen_supplier_reconciliation_xlsx(path: Path) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl 缺失,跳过 .xlsx", file=sys.stderr)
        return 0

    wb = Workbook()
    # Sheet 1: 月度对账
    ws1 = wb.active
    ws1.title = "月度对账"
    ws1["A1"] = "宜兴市锦泰耐火材料 · 供应商对账单 (演示样本)"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:F1")
    headers = ["供应商", "月份", "出库笔数", "入库金额 (元)", "应付余额 (元)", "账期"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2c3e50")
        c.alignment = Alignment(horizontal="center")
    rows = [
        ("山东中铝物资", "2026-04", 6, 432_000, 86_400, "60 天"),
        ("山东中铝物资", "2026-05", 5, 528_000, 105_600, "60 天"),
        ("淄博耐材", "2026-05", 3, 168_000, 33_600, "30 天"),
        ("郑州磷酸盐", "2026-05", 2, 76_000, 15_200, "45 天"),
    ]
    for r, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            ws1.cell(row=r, column=c, value=v)
    for col in range(1, 7):
        ws1.column_dimensions[ws1.cell(row=3, column=col).column_letter].width = 18

    # Sheet 2: 货款明细
    ws2 = wb.create_sheet("货款明细")
    ws2.append(["PO 编号", "入库日期", "金额 (元)", "已付 (元)", "未付 (元)", "状态"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.append(["PO-2026-005", "2026-04-08", 96_000, 96_000, 0, "已结清"])
    ws2.append(["PO-2026-008", "2026-04-22", 124_000, 0, 124_000, "待支付"])
    ws2.append(["PO-2026-009", "2026-05-19", 96_000, 0, 96_000, "未到期"])

    # Sheet 3: 备注
    ws3 = wb.create_sheet("备注")
    ws3["A1"] = "本文档为锦泰 demo 演示样本"
    ws3["A1"].font = Font(bold=True, color="C0392B")
    ws3["A3"] = "上传到 backend mode 后,后端 DemoMockProvider 会基于文件名 + 大小派生 IssueVoucher 候选。"
    ws3["A4"] = "真实环境下走 ClaudeProvider (ANTHROPIC_API_KEY 设置时) 真 AI 抽取。"

    wb.save(path)
    print(f"  ✓ {path.relative_to(REPO_ROOT)}  ({path.stat().st_size} bytes)")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
